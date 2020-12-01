from openpyxl import Workbook, load_workbook
import TestData

# workbook=Workbook()
# sheet=workbook.active
# sheet.title="test"
# sheet["A1"]="Hello"
# sheet["B1"]="World!"
# workbook.save(filename="ExcelTest.xlsx")
workbook=load_workbook(filename="sample.xlsx")
sheet=workbook.active
print(workbook.sheetnames)
print(sheet["A1"].value)
print(sheet.cell(row=1, column=1).value)
for i in sheet.iter_rows(min_row=1,max_row=3,values_only=True):
    print(i)
